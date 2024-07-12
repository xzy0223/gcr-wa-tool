import csv
from openpyxl import load_workbook

'''
这个脚本的主要作用是从客户导出的trust advisor report中提取出ta check和对应description的mapping表
用于丰富TAM导出的report，让客户可以更好的理解每个TA check
'''

workbook = load_workbook('./all.xlsx')

sheet_names = workbook.sheetnames

with open('/home/ec2-user/wa-tool/ta-check-desc.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(['ta_check', 'description'])

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        ta_check = sheet['A1'].value
        description = sheet['A3'].value
        writer.writerow([ta_check, description])


